from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sqlite3
import subprocess
import tempfile
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pdfplumber
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import HexColor, white
from reportlab.pdfgen import canvas

PdfBBox = tuple[float, float, float, float]


DIMENSION_RE = re.compile(
    r"""
    ^
    (?:
      (?:[0-9]+x\s*)?
      (?:
        \(?\s*(?:\+/-|\u00b1)?\s*[drms]?\s*(?:[0-9]+(?:[.,][0-9]+)?|[.,][0-9]+)\s*\)?
        |
        \(?\s*[0-9]+/[0-9]+\s*\)?
        |
        \(?\s*[0-9]+\s*-\s*[0-9]+\s*\)?
      )
      (?:\s*(?:mm|cm|in|deg|degree|degrees|grados|"|\u00b0)?)?
      (?:\s*(?:\+/-|\u00b1|\+|-)\s*(?:[0-9]+(?:[.,][0-9]+)?|[.,][0-9]+))?
      (?:\s*/\s*(?:\+|-)?\s*[0-9]+(?:[.,][0-9]+)?)?
    )
    $
    """,
    re.IGNORECASE | re.VERBOSE,
)

NOT_DIMENSION_RE = re.compile(
    r"(?i)^(?:rev|revision|sheet|page|date|scale|material|qty|cantidad|cliente|client)$"
)
NON_DIMENSION_LINE_RE = re.compile(
    r"(?i)\b(?:rev|revision|sheet|page|date|scale|material|qty|cantidad|cliente|client)\b"
)
GENERAL_TOLERANCE_LINE_RE = re.compile(
    r"(?i)\b(?:"
    r"unless\s+otherwise\s+specified|"
    r"tolerances?|"
    r"dimensions?\s+are\s+in|"
    r"angles?|"
    r"surface\s+finish|"
    r"break\s+sharp\s+corners?|"
    r"max\s+surface\s+finish|"
    r"third\s+angle\s+projection"
    r")\b"
)
GENERAL_TOLERANCE_NOTATION_RE = re.compile(
    r"(?i)(?:\.[Xx]{1,4}|(?<!\d)[Xx]{2,4})\s*(?:\+/-|[+]\s*/\s*-|[Â±±]|[+]\s*[-])\s*\.?\d+"
)

NOTE_CONTEXT_LINE_RE = re.compile(
    r"(?i)\b(?:notes?|material|finish|deburr|break\s+edges?|remove\s+burrs?|surface|"
    r"plating|coating|anodize|passivate|heat\s+treat|unless\s+otherwise)\b"
)
HEADER_STATUS_LINE_RE = re.compile(
    r"(?i)\b(?:state|released|date|effectivity|upon\s+release|gmt)\b"
)


@dataclass
class DimensionCandidate:
    number: int
    page: int
    text: str
    x: float
    y: float
    width: float
    height: float
    confidence: float
    reason: str


@dataclass(frozen=True)
class AnalysisProfile:
    name: str
    include_tables: bool = False
    use_rotated: bool = True
    use_stacked: bool = True
    use_ocr: bool = False
    min_confidence: float = 0.0
    allow_plain_numbers: bool = True
    exclude_title_block: bool = True
    right_exclusion_ratio: float = 0.0
    description: str = ""


ANALYSIS_PROFILES: dict[str, AnalysisProfile] = {
    "standard": AnalysisProfile(
        name="standard",
        description="PDF vectorial normal; excluye tablas, cajetines y notas.",
    ),
    "conservative": AnalysisProfile(
        name="conservative",
        min_confidence=0.74,
        allow_plain_numbers=False,
        description="Mas estricto; prioriza cotas con marcas tecnicas, unidades o tolerancias.",
    ),
    "permissive": AnalysisProfile(
        name="permissive",
        include_tables=True,
        min_confidence=0.55,
        description="Mas amplio para planos con formatos no estandar o cotas en zonas tipo tabla.",
    ),
    "vertical_side": AnalysisProfile(
        name="vertical_side",
        min_confidence=0.70,
        right_exclusion_ratio=0.30,
        description="Planos en vertical con cajetin lateral derecho; bloquea esa franja.",
    ),
    "ocr": AnalysisProfile(
        name="ocr",
        use_ocr=True,
        description="Fuerza OCR para planos escaneados como imagen.",
    ),
}

DEFAULT_ANALYSIS_STRATEGY = "auto"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_text(text: str) -> str:
    return (
        text.strip()
        .replace(",", ".")
        .replace("\uf060", "\u00b1")
        .replace("\uf0b1", "\u00b1")
        .replace("\uf06e", "\u2300")
        .replace("\uf06d", "M")
        .replace("\uf078", "")
        .replace("\uf06a", "")
        .replace("\u00d8", "D")
        .replace("\u00f8", "D")
        .replace("\u2300", "D")
        .replace("\u00b0", "deg")
    )


def parse_decimal_number(raw_value: str) -> float | None:
    value = normalize_text(raw_value).strip()
    if not value:
        return None
    if value.startswith("."):
        value = f"0{value}"
    if value.startswith("-."):
        value = value.replace("-.", "-0.", 1)
    if value.startswith("+."):
        value = value.replace("+.", "+0.", 1)
    try:
        return float(value)
    except ValueError:
        return None


def decimal_places_from_token(token: str) -> int:
    clean = normalize_text(token).strip()
    if "." not in clean:
        return 0
    return len(re.sub(r"[^0-9]", "", clean.split(".", 1)[1]))


def tolerance_by_decimals(decimal_places: int) -> float | None:
    if decimal_places == 2:
        return 0.01
    if decimal_places == 3:
        return 0.005
    if decimal_places >= 4:
        return 0.001
    return None


def is_thread_callout(text: str) -> bool:
    clean = normalize_text(text).upper()
    return bool(
        (
            re.search(r"\b\d+\s*-\s*\d+\b", clean)
            and re.search(r"\b(?:UNC|UNF|UNEF|THREAD|THD)\b", clean)
        )
        or re.search(r"\bM\d+(?:\.\d+)?\s*X\s*\d+(?:\.\d+)?(?:\s*-\s*\d+[A-Z])?", clean)
    )


def is_date_like_text(text: str) -> bool:
    clean = normalize_text(text).strip()
    return bool(
        re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", clean)
        or re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", clean)
        or re.fullmatch(r"\d{1,2}[.-]\d{1,2}[.-]\d{2,4}", clean)
    )


def line_has_date_fragment(text: str) -> bool:
    clean = normalize_text(text)
    return bool(
        re.search(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", clean)
        or re.search(r"\b\d{1,2}\s+\d{1,2}\s+\d{2,4}\b", clean)
        or re.search(r"\b\d{4}-\d{1,2}-\d{1,2}\b", clean)
    )


def inspection_nominal_text(text: str) -> str:
    clean = normalize_text(text).strip()
    clean = re.sub(r"^\((.*)\)$", r"\1", clean).strip()
    clean = re.sub(r"(?i)^\s*\d+\s*X\s+", "", clean)
    clean = re.sub(r"(?i)^\s*\d+\s*X(?=[RMS]?\s*(?:\d|\.))", "", clean)
    return clean.strip()


def tolerance_info(text: str) -> dict[str, Any]:
    clean = inspection_nominal_text(text)
    unit_match = re.search(r"(?i)\b(mm|cm|in|deg|grados)\b|\"", clean)
    unit = unit_match.group(0).replace('"', "in").lower() if unit_match else ""

    explicit_plus: float | None = None
    explicit_minus: float | None = None
    plus_minus = re.search(r"(?:\+/-|±)\s*([+-]?(?:\d+(?:\.\d+)?|\.\d+))", clean)
    if plus_minus:
        explicit = parse_decimal_number(plus_minus.group(1))
        if explicit is not None:
            explicit_plus = abs(explicit)
            explicit_minus = -abs(explicit)
    else:
        explicit_pairs = re.findall(r"([+-])\s*((?:\d+(?:\.\d+)?|\.\d+))", clean)
        for sign, value_text in explicit_pairs:
            value = parse_decimal_number(value_text)
            if value is None:
                continue
            if sign == "+" and explicit_plus is None:
                explicit_plus = abs(value)
            elif sign == "-" and explicit_minus is None:
                explicit_minus = -abs(value)

    radius_match = re.search(r"(?i)\bR\s*((?:\d+\.\d+|\.\d+|\d+))", clean)
    diameter_match = re.search(r"(?i)\bD\s*((?:\d+\.\d+|\.\d+|\d+))", clean)
    value_tokens = [radius_match.group(1)] if radius_match else [diameter_match.group(1)] if diameter_match else re.findall(r"(?<![A-Za-z])(?:\d+\.\d+|\.\d+|\d+)(?![A-Za-z])", clean)
    nominal_token = ""
    nominal: float | None = None
    for token in value_tokens:
        previous = clean[max(0, clean.find(token) - 2) : clean.find(token)]
        if "+" in previous or "-" in previous or "+/-" in previous:
            continue
        nominal = parse_decimal_number(token)
        nominal_token = token
        if nominal is not None:
            break

    decimals = decimal_places_from_token(nominal_token) if nominal_token else 0
    source = ""
    tol_plus = explicit_plus
    tol_minus = explicit_minus
    if tol_plus is not None or tol_minus is not None:
        source = "explicita en cota"
        tol_plus = tol_plus if tol_plus is not None else 0.0
        tol_minus = tol_minus if tol_minus is not None else 0.0
    else:
        general = tolerance_by_decimals(decimals)
        if general is not None:
            tol_plus = general
            tol_minus = -general
            source = f"general {decimals} decimales"

    minimum = nominal + tol_minus if nominal is not None and tol_minus is not None else None
    maximum = nominal + tol_plus if nominal is not None and tol_plus is not None else None
    return {
        "nominal": nominal,
        "unit": unit,
        "decimals": decimals,
        "tol_plus": tol_plus,
        "tol_minus": tol_minus,
        "minimum": minimum,
        "maximum": maximum,
        "source": source or "sin tolerancia asignada",
    }


def looks_like_dimension(
    raw_text: str,
    line_text: str = "",
    neighbor_text: str = "",
) -> tuple[bool, float, str]:
    
    text = normalize_text(raw_text)
    compact = re.sub(r"\s+", "", text)

    if is_date_like_text(text):
        return False, 0.0, "ignored date"

    if not compact or NOT_DIMENSION_RE.match(compact):
        return False, 0.0, "ignored label"

    if not re.search(r"\d", compact):
        return False, 0.0, "no digits"

    line_tokens = [token for token in re.split(r"\s+", line_text.strip()) if token]
    context_text = normalize_text(" ".join(part for part in [line_text, neighbor_text] if part))
    line_has_unit_nearby = bool(
        re.search(r"(?i)\b(mm|cm|in|deg|grados|aprox|approx)\b|\"", context_text)
    )

    if compact in {"99", "01", "0.0", "00"}:
        return False, 0.0, "ignored ocr artifact"
    if re.fullmatch(r"\(?[+-]?(?:0+(?:[.,]0+)?|[.,]0+)\)?", compact):
        return False, 0.0, "ignored origin zero"
    if re.match(r"^[0-9]{3}$", compact) and compact not in {"110"} and not line_has_unit_nearby:
        return False, 0.0, "ignored unlikely ocr integer"

    has_strong_marker = bool(
        re.search(r"(?i)(^\(?[rdm]|x|\+/-|\u00b1|\+|-|/|mm|cm|in|deg|grados|\")", compact)
    )
    if re.match(r"^[0-9]$", compact) and not has_strong_marker and not line_has_unit_nearby:
        return False, 0.0, "ignored single digit"

    if line_text and NON_DIMENSION_LINE_RE.search(line_text) and not has_strong_marker:
        return False, 0.0, "ignored title block or note line"
    if line_text and "/" in line_text and re.search(r"(?i)\b(?:rev|revision|date|approved|history)\b", line_text):
        return False, 0.0, "ignored revision date fragment"
    if line_text and line_has_date_fragment(line_text) and re.search(
        r"(?i)\b(?:rev|revision|date|approved|history|add|was|clarified|release|initial|description)\b",
        line_text,
    ):
        return False, 0.0, "ignored revision history row"
    if re.match(r"^\d{1,2}$", compact) and re.search(r"\b\d{4}\b", line_text) and not line_has_unit_nearby:
        return False, 0.0, "ignored date number fragment"
    if line_text and GENERAL_TOLERANCE_LINE_RE.search(line_text):
        return False, 0.0, "ignored general tolerance block"
    if line_text and GENERAL_TOLERANCE_NOTATION_RE.search(line_text):
        return False, 0.0, "ignored general tolerance notation"
    if line_text and NOTE_CONTEXT_LINE_RE.search(line_text):
        return False, 0.0, "ignored note line"
    if line_text and HEADER_STATUS_LINE_RE.search(line_text):
        return False, 0.0, "ignored drawing header line"

    is_plain_number = bool(re.match(r"^[0-9]+(?:[.,][0-9]+)?$", compact))
    if is_plain_number and len(line_tokens) > 12 and not line_has_unit_nearby:
        return False, 0.0, "plain number inside note"

    if DIMENSION_RE.match(compact):
        if re.match(r"^[0-9]{4,}$", compact):
            return False, 0.0, "plain long integer"

        confidence = 0.74
        reason = "dimension pattern"
        if re.search(r"(?i)(mm|cm|in|deg|grados|\")", compact):
            confidence += 0.08
            reason += " with unit"
        if re.search(r"(?i)(^\(?[rdm]|x|\+/-|\u00b1|\+|-|/)", compact):
            confidence += 0.08
            reason += " with technical marker"
        return True, min(confidence, 0.96), reason

    if is_plain_number:
        confidence = 0.72 if line_has_unit_nearby else 0.62
        reason = "plain numeric value with nearby unit" if line_has_unit_nearby else "plain numeric value"
        return True, confidence, reason

    if is_thread_callout(text):
        return True, 0.9, "thread callout"

    return False, 0.0, "not a dimension"


def word_center(word: dict[str, Any]) -> tuple[float, float]:
    return (
        (float(word["x0"]) + float(word["x1"])) / 2,
        (float(word["top"]) + float(word["bottom"])) / 2,
    )


def point_inside_bbox(x: float, y: float, bbox: PdfBBox, padding: float = 2.0) -> bool:
    x0, top, x1, bottom = bbox
    return (x0 - padding) <= x <= (x1 + padding) and (top - padding) <= y <= (bottom + padding)


def word_inside_any_bbox(word: dict[str, Any], bboxes: list[PdfBBox]) -> bool:
    x, y = word_center(word)
    return any(point_inside_bbox(x, y, bbox) for bbox in bboxes)


def detected_table_bboxes(page: pdfplumber.page.Page) -> list[PdfBBox]:
    bboxes: list[PdfBBox] = []
    try:
        tables = page.find_tables()
    except Exception:
        return bboxes

    page_area = float(page.width) * float(page.height)
    for table in tables:
        x0, top, x1, bottom = [float(value) for value in table.bbox]
        width = x1 - x0
        height = bottom - top
        if width <= 0 or height <= 0:
            continue

        area_ratio = (width * height) / page_area
        page_width_ratio = width / float(page.width)
        page_height_ratio = height / float(page.height)

        # pdfplumber can mistake the drawing border/grid for one huge table.
        # If we exclude that box, every dimension inside the drawing disappears.
        if area_ratio > 0.60 or (page_width_ratio > 0.85 and page_height_ratio > 0.75):
            continue

        # Avoid treating a small dimension callout box as a table. Real title
        # blocks and reference tables usually occupy a meaningful page area.
        if area_ratio < 0.015:
            continue

        near_page_edge = (
            bottom >= float(page.height) * 0.72
            or x0 <= float(page.width) * 0.06
            or x1 >= float(page.width) * 0.94
        )
        if not near_page_edge:
            continue

        bboxes.append((x0, top, x1, bottom))

    return bboxes


def detected_title_block_bboxes(words: list[dict[str, Any]], page_width: float, page_height: float) -> list[PdfBBox]:
    if not words:
        return []

    keyword_re = re.compile(
        r"(?i)\b(?:"
        r"date|drawn|checked|approved|revision|rev|sheet|of|scale|title|description|"
        r"material|finish|cage|code|dwg|drawing|part|number|third\s+angle|projection|"
        r"tolerances?|dimensions?|unless|specified|angular|decimal|fraction|notes?"
        r")\b"
    )
    page_area = page_width * page_height
    boxes: list[PdfBBox] = []

    edge_regions = [
        ("right", [word for word in words if float(word.get("x0", 0)) >= page_width * 0.68]),
        ("bottom", [word for word in words if float(word.get("top", 0)) >= page_height * 0.72]),
        ("left", [word for word in words if float(word.get("x1", 0)) <= page_width * 0.20]),
    ]

    for region, region_words in edge_regions:
        if len(region_words) < 8:
            continue
        keyword_words = [
            word
            for word in region_words
            if keyword_re.search(normalize_text(str(word.get("text", ""))))
        ]
        if len(keyword_words) < 3:
            continue

        if region == "right":
            keyword_clusters: list[list[dict[str, Any]]] = []
            for word in sorted(keyword_words, key=lambda item: float(item.get("top", 0))):
                if not keyword_clusters:
                    keyword_clusters.append([word])
                    continue
                previous_bottom = max(float(item.get("bottom", 0)) for item in keyword_clusters[-1])
                if float(word.get("top", 0)) - previous_bottom > 72:
                    keyword_clusters.append([word])
                else:
                    keyword_clusters[-1].append(word)

            for keyword_cluster in keyword_clusters:
                if len(keyword_cluster) < 3:
                    continue
                keyword_union = union_pdf_words(keyword_cluster)
                top_limit = max(0.0, keyword_union["y"] - 44.0)
                bottom_limit = min(page_height, keyword_union["y"] + keyword_union["height"] + 44.0)
                block_words = [
                    word
                    for word in region_words
                    if top_limit
                    <= (float(word.get("top", 0)) + float(word.get("bottom", 0))) / 2
                    <= bottom_limit
                ]
                if len(block_words) < 6:
                    continue
                block_union = union_pdf_words(block_words)
                if block_union["width"] > page_width * 0.38:
                    continue
                boxes.append(
                    (
                        max(0.0, min(float(word.get("x0", 0)) for word in keyword_cluster) - 26.0),
                        max(0.0, block_union["y"] - 12.0),
                        page_width,
                        min(page_height, block_union["y"] + block_union["height"] + 12.0),
                    )
                )
            continue

        union = union_pdf_words(region_words)
        width = union["width"]
        height = union["height"]
        if width <= 0 or height <= 0:
            continue

        area_ratio = (width * height) / page_area
        vertical_strip = region == "right" and height >= page_height * 0.32 and width <= page_width * 0.36
        bottom_block = region == "bottom" and width >= page_width * 0.32 and height <= page_height * 0.34
        left_revision = region == "left" and height >= page_height * 0.20 and width <= page_width * 0.24
        if not (vertical_strip or bottom_block or left_revision or area_ratio >= 0.05):
            continue

        if region == "right":
            x0 = max(0.0, min(float(word.get("x0", 0)) for word in keyword_words) - 26.0)
            top = max(0.0, union["y"] - 18.0)
            x1 = page_width
            bottom = min(page_height, union["y"] + union["height"] + 18.0)
        elif region == "bottom":
            x0 = max(0.0, union["x"] - 18.0)
            top = max(0.0, min(float(word.get("top", 0)) for word in keyword_words) - 18.0)
            x1 = min(page_width, union["x"] + union["width"] + 18.0)
            bottom = page_height
        else:
            x0 = 0.0
            top = max(0.0, union["y"] - 18.0)
            x1 = min(page_width, union["x"] + union["width"] + 18.0)
            bottom = min(page_height, union["y"] + union["height"] + 18.0)

        boxes.append((x0, top, x1, bottom))

    return boxes


def profile_excluded_bboxes(profile: AnalysisProfile, page_width: float, page_height: float) -> list[PdfBBox]:
    boxes: list[PdfBBox] = []
    if profile.right_exclusion_ratio > 0:
        left = page_width * (1.0 - profile.right_exclusion_ratio)
        boxes.append((left, 0.0, page_width, page_height))
    return boxes


def detected_note_bboxes(words: list[dict[str, Any]], page_width: float, page_height: float) -> list[PdfBBox]:
    if not words:
        return []

    lines: list[list[dict[str, Any]]] = []
    for word in sorted(words, key=lambda item: (float(item.get("top", 0)), float(item.get("x0", 0)))):
        if not lines or abs(float(lines[-1][0].get("top", 0)) - float(word.get("top", 0))) >= 4:
            lines.append([word])
        else:
            lines[-1].append(word)

    boxes: list[PdfBBox] = []
    for index, line_words in enumerate(lines):
        line_text = normalize_text(" ".join(str(word.get("text", "")) for word in line_words))
        line_box = union_pdf_words(line_words)
        line_is_notes_header = bool(re.search(r"(?i)\bnotes?\s*:?\b", line_text))
        line_is_numbered_note = bool(
            re.search(r"^\s*\d+\s*[\.)-]", line_text)
            and NOTE_CONTEXT_LINE_RE.search(line_text)
        )
        if not line_is_notes_header and not line_is_numbered_note:
            continue

        block_words = list(line_words)
        left_limit = max(0.0, line_box["x"] - 30.0)
        right_limit = min(page_width, line_box["x"] + max(line_box["width"], page_width * 0.26) + 180.0)
        bottom_limit = min(page_height, line_box["y"] + max(page_height * 0.30, 180.0))
        for extra_line in lines[index + 1 : index + 10]:
            extra_text = normalize_text(" ".join(str(word.get("text", "")) for word in extra_line))
            extra_box = union_pdf_words(extra_line)
            extra_center_x = extra_box["x"] + extra_box["width"] / 2
            if extra_box["y"] > bottom_limit:
                break
            if left_limit <= extra_center_x <= right_limit and (
                NOTE_CONTEXT_LINE_RE.search(extra_text)
                or re.search(r"^\s*\d+\s*[\.)-]", extra_text)
            ):
                block_words.extend(extra_line)

        union = union_pdf_words(block_words)
        boxes.append(
            (
                max(0.0, union["x"] - 14.0),
                max(0.0, union["y"] - 10.0),
                min(page_width, union["x"] + union["width"] + 14.0),
                min(page_height, union["y"] + union["height"] + 10.0),
            )
        )

    return boxes


def detected_general_tolerance_bboxes(words: list[dict[str, Any]], page_width: float, page_height: float) -> list[PdfBBox]:
    boxes: list[PdfBBox] = []
    if not words:
        return boxes

    def split_horizontal_clusters(line_words: list[dict[str, Any]], gap_limit: float = 42.0) -> list[list[dict[str, Any]]]:
        clusters: list[list[dict[str, Any]]] = []
        for word in sorted(line_words, key=lambda item: float(item.get("x0", 0))):
            if not clusters:
                clusters.append([word])
                continue
            previous = clusters[-1][-1]
            gap = float(word.get("x0", 0)) - float(previous.get("x1", 0))
            if gap > gap_limit:
                clusters.append([word])
            else:
                clusters[-1].append(word)
        return clusters

    lines: list[list[dict[str, Any]]] = []
    for word in sorted(words, key=lambda item: (float(item.get("top", 0)), float(item.get("x0", 0)))):
        if not lines or abs(float(lines[-1][0].get("top", 0)) - float(word.get("top", 0))) >= 4:
            lines.append([word])
        else:
            lines[-1].append(word)

    for index, line_words in enumerate(lines):
        trigger_clusters = []
        for cluster in split_horizontal_clusters(line_words):
            cluster_text = normalize_text(" ".join(str(word.get("text", "")) for word in cluster))
            if GENERAL_TOLERANCE_LINE_RE.search(cluster_text):
                trigger_clusters.append(cluster)
        if not trigger_clusters:
            continue

        for trigger_words in trigger_clusters:
            line_text = normalize_text(" ".join(str(word.get("text", "")) for word in trigger_words))
            line_box = union_pdf_words(trigger_words)
            line_is_title_block_context = (
                line_box["y"] >= page_height * 0.58
                or line_box["x"] <= page_width * 0.08
                or "UNLESS OTHERWISE SPECIFIED" in line_text.upper()
                or "DIMENSIONS ARE IN" in line_text.upper()
                or "TOLERANCES UNLESS" in line_text.upper()
            )
            if not line_is_title_block_context:
                continue

            block_words = list(trigger_words)
            left_limit = max(0.0, line_box["x"] - 35.0)
            right_limit = min(page_width, line_box["x"] + line_box["width"] + 140.0)
            for extra_line in lines[index + 1 : index + 8]:
                if not extra_line:
                    continue
                first_top = float(extra_line[0].get("top", 0))
                if first_top > page_height * 0.97:
                    break
                for cluster in split_horizontal_clusters(extra_line):
                    cluster_box = union_pdf_words(cluster)
                    cluster_center_x = cluster_box["x"] + cluster_box["width"] / 2
                    if left_limit <= cluster_center_x <= right_limit:
                        block_words.extend(cluster)

            union = union_pdf_words(block_words)
            x0 = max(0.0, union["x"] - 10)
            top = max(0.0, union["y"] - 6)
            x1 = min(page_width, union["x"] + union["width"] + 10)
            bottom = min(page_height, union["y"] + union["height"] + 6)
            if bottom >= page_height * 0.55 or x0 <= page_width * 0.35:
                boxes.append((x0, top, x1, bottom))

    return boxes


def detected_revision_history_bboxes(words: list[dict[str, Any]], page_width: float, page_height: float) -> list[PdfBBox]:
    boxes: list[PdfBBox] = []
    if not words:
        return boxes

    lines: list[list[dict[str, Any]]] = []
    for word in sorted(words, key=lambda item: (float(item.get("top", 0)), float(item.get("x0", 0)))):
        if not lines or abs(float(lines[-1][0].get("top", 0)) - float(word.get("top", 0))) >= 4:
            lines.append([word])
        else:
            lines[-1].append(word)

    for index, line_words in enumerate(lines):
        line_text = normalize_text(" ".join(str(word.get("text", "")) for word in line_words)).upper()
        line_box = union_pdf_words(line_words)
        if "REVISION" not in line_text and "REV." not in line_text:
            continue
        if line_box["y"] > page_height * 0.25 or line_box["x"] < page_width * 0.45:
            continue

        block_words = list(line_words)
        for extra_line in lines[index + 1 : index + 6]:
            if not extra_line:
                continue
            extra_box = union_pdf_words(extra_line)
            if extra_box["y"] > page_height * 0.28:
                break
            if extra_box["x"] >= page_width * 0.42:
                block_words.extend(extra_line)

        union = union_pdf_words(block_words)
        boxes.append(
            (
                max(0.0, union["x"] - 12),
                max(0.0, union["y"] - 8),
                min(page_width, union["x"] + union["width"] + 12),
                min(page_height, union["y"] + union["height"] + 8),
            )
        )

    return boxes


def nearby_line_words(words: list[dict[str, Any]], word: dict[str, Any]) -> list[dict[str, Any]]:
    top = float(word.get("top", 0))
    peers = [peer for peer in words if abs(float(peer.get("top", 0)) - top) < 3]
    return sorted(peers, key=lambda item: float(item.get("x0", 0)))


def nearby_context(line_words: list[dict[str, Any]], word: dict[str, Any]) -> str:
    try:
        index = line_words.index(word)
    except ValueError:
        return ""

    start = max(0, index - 1)
    end = min(len(line_words), index + 2)
    return " ".join(str(peer.get("text", "")) for peer in line_words[start:end])


def word_bbox_from_pdf_word(word: dict[str, Any]) -> dict[str, float]:
    return {
        "x": float(word["x0"]),
        "y": float(word["top"]),
        "width": float(word["x1"] - word["x0"]),
        "height": float(word["bottom"] - word["top"]),
    }


def union_pdf_words(words: list[dict[str, Any]]) -> dict[str, float]:
    xs = [float(word["x0"]) for word in words]
    ys = [float(word["top"]) for word in words]
    rights = [float(word["x1"]) for word in words]
    bottoms = [float(word["bottom"]) for word in words]
    return {
        "x": min(xs),
        "y": min(ys),
        "width": max(rights) - min(xs),
        "height": max(bottoms) - min(ys),
    }


def reverse_rotated_text(text: str) -> str:
    return normalize_text(text[::-1])


def extract_rotated_word_candidates(
    words: list[dict[str, Any]],
    page_index: int,
    page_width: float,
    page_height: float,
    table_bboxes: list[PdfBBox] | None = None,
) -> list[dict[str, Any]]:
    table_bboxes = table_bboxes or []
    rotated_words = [
        word
        for word in words
        if not bool(word.get("upright", True)) and not word_inside_any_bbox(word, table_bboxes)
    ]
    rotated_words.sort(key=lambda item: (float(item.get("x0", 0)), float(item.get("top", 0))))

    candidates: list[dict[str, Any]] = []
    consumed: set[int] = set()

    def same_rotated_column(a: dict[str, Any], b: dict[str, Any], tolerance: float = 12.0) -> bool:
        ax = (float(a.get("x0", 0)) + float(a.get("x1", 0))) / 2
        bx = (float(b.get("x0", 0)) + float(b.get("x1", 0))) / 2
        return abs(ax - bx) <= tolerance

    def rotated_word_gap(a: dict[str, Any], b: dict[str, Any]) -> float:
        return min(
            abs(float(a.get("top", 0)) - float(b.get("bottom", 0))),
            abs(float(b.get("top", 0)) - float(a.get("bottom", 0))),
            abs(float(a.get("top", 0)) - float(b.get("top", 0))),
            abs(float(a.get("bottom", 0)) - float(b.get("bottom", 0))),
        )

    for index, word in enumerate(rotated_words):
        if index in consumed:
            continue

        text = reverse_rotated_text(str(word.get("text", ""))).strip()
        if not text:
            continue

        phrase_words = [word]
        phrase_text = text
        if is_prefix_dimension_token(text):
            best_value_index: int | None = None
            best_gap = 999.0
            for other_index, other in enumerate(rotated_words):
                if other_index == index or other_index in consumed:
                    continue
                if not same_rotated_column(other, word):
                    continue

                other_text = reverse_rotated_text(str(other.get("text", ""))).strip()
                ok, _, _ = looks_like_dimension(other_text)
                if not ok:
                    continue

                gap = rotated_word_gap(word, other)
                if gap < best_gap and gap <= 18:
                    best_gap = gap
                    best_value_index = other_index

            if best_value_index is not None:
                phrase_words.append(rotated_words[best_value_index])
                phrase_text = f"{text} {reverse_rotated_text(str(rotated_words[best_value_index].get('text', ''))).strip()}"
                consumed.add(best_value_index)
        else:
            best_suffix_index: int | None = None
            best_suffix_gap = 999.0
            for other_index, other in enumerate(rotated_words):
                if other_index == index or other_index in consumed:
                    continue
                if not same_rotated_column(other, word):
                    continue

                other_text = reverse_rotated_text(str(other.get("text", ""))).strip()
                if not is_suffix_dimension_token(other_text):
                    continue

                gap = rotated_word_gap(word, other)
                if gap < best_suffix_gap and gap <= 24:
                    best_suffix_gap = gap
                    best_suffix_index = other_index

            if best_suffix_index is not None:
                phrase_words.append(rotated_words[best_suffix_index])
                phrase_text = f"{text} {reverse_rotated_text(str(rotated_words[best_suffix_index].get('text', ''))).strip()}"
                consumed.add(best_suffix_index)

            best_prefix_index: int | None = None
            best_gap = 999.0
            for other_index, other in enumerate(rotated_words):
                if other_index == index or other_index in consumed:
                    continue
                if not same_rotated_column(other, word):
                    continue

                other_text = reverse_rotated_text(str(other.get("text", ""))).strip()
                if not is_prefix_dimension_token(other_text):
                    continue

                gap = rotated_word_gap(word, other)
                if gap < best_gap and gap <= 18:
                    best_gap = gap
                    best_prefix_index = other_index

            if best_prefix_index is not None:
                phrase_words.append(rotated_words[best_prefix_index])
                phrase_text = f"{reverse_rotated_text(str(rotated_words[best_prefix_index].get('text', ''))).strip()} {text}"
                consumed.add(best_prefix_index)

        ok, confidence, reason = looks_like_dimension(phrase_text, phrase_text, phrase_text)
        if not ok:
            continue

        box = union_pdf_words(phrase_words)
        in_title_block = box["x"] >= page_width * 0.55 and box["y"] >= page_height * 0.86
        if in_title_block:
            continue

        consumed.add(index)
        candidates.append(
            {
                "page": page_index,
                "text": phrase_text,
                **box,
                "confidence": round(min(confidence + 0.04, 0.96), 3),
                "reason": f"rotated {reason}",
            }
        )

    return candidates


def extract_stacked_autocad_decimal_candidates(
    chars: list[dict[str, Any]],
    page_index: int,
    excluded_bboxes: list[PdfBBox] | None = None,
) -> list[dict[str, Any]]:
    excluded_bboxes = excluded_bboxes or []
    usable: list[dict[str, Any]] = []
    for char in chars:
        text = normalize_text(str(char.get("text", ""))).strip()
        if text not in {".", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9"}:
            continue
        x0 = float(char.get("x0", 0))
        top = float(char.get("top", 0))
        x1 = float(char.get("x1", x0))
        bottom = float(char.get("bottom", top))
        if word_inside_any_bbox({"x0": x0, "x1": x1, "top": top, "bottom": bottom}, excluded_bboxes):
            continue

        width = x1 - x0
        height = bottom - top
        # AutoCAD sometimes exports rotated dimensions as large individual
        # upright glyphs stacked in a column. Small title-block tolerances are
        # intentionally below this threshold.
        if width < 18 or height < 8:
            continue
        usable.append({"text": text, "x0": x0, "top": top, "x1": x1, "bottom": bottom})

    columns: list[list[dict[str, Any]]] = []
    for char in sorted(usable, key=lambda item: ((float(item["x0"]) + float(item["x1"])) / 2, float(item["top"]))):
        center_x = (float(char["x0"]) + float(char["x1"])) / 2
        placed = False
        for column in columns:
            column_center = sum((float(item["x0"]) + float(item["x1"])) / 2 for item in column) / len(column)
            if abs(center_x - column_center) <= 6:
                column.append(char)
                placed = True
                break
        if not placed:
            columns.append([char])

    candidates: list[dict[str, Any]] = []
    column_groups: list[list[dict[str, Any]]] = []
    for column in columns:
        ordered_column = sorted(column, key=lambda item: float(item["top"]))
        current: list[dict[str, Any]] = []
        for char in ordered_column:
            if current and float(char["top"]) - float(current[-1]["bottom"]) > 48:
                column_groups.append(current)
                current = []
            current.append(char)
        if current:
            column_groups.append(current)

    for column in column_groups:
        if len(column) not in {3, 4}:
            continue
        ordered = sorted(column, key=lambda item: float(item["top"]))
        texts = [str(item["text"]) for item in ordered]
        if "." not in texts or sum(1 for text in texts if text.isdigit()) != len(texts) - 1:
            continue

        value = "".join(reversed(texts))
        if not re.match(r"^(?:\.\d{2}|\d\.\d{2})$", value):
            continue
        if value.endswith(".00") or value == ".00":
            continue

        box = union_pdf_words(ordered)
        candidates.append(
            {
                "page": page_index,
                "text": value,
                "x": box["x"],
                "y": box["y"],
                "width": box["width"],
                "height": box["height"],
                "confidence": 0.78,
                "reason": "stacked AutoCAD decimal dimension",
            }
        )

    return candidates


def is_prefix_dimension_token(text: str) -> bool:
    compact = normalize_text(text).upper()
    return bool(re.match(r"^[0-9]+X$", compact))


def is_suffix_dimension_token(text: str) -> bool:
    compact = normalize_text(text).upper().strip(",")
    return bool(
        compact
        in {
            "THRU",
            "ALL",
            "UNC",
            "UNF",
            "UNEF",
            "THREAD",
            "THD",
            "MINOR",
            "MAJOR",
            "DIAMETER",
            "NEAR",
            "FAR",
            "SIDE",
            "WALL",
            "TYP",
            "MM",
            "CM",
            "IN",
            "APROX",
            "APROX.",
            "APPROX",
            "APPROX.",
            "EQ",
            "SP",
            "PLCS",
            "PLACES",
            "DEG",
            "°",
            "+",
            "-",
        }
        or re.match(r"^\([0-9]+\)$", compact)
        or re.match(r"^[0-9]+(?!X$)[A-Z]$", compact)
    )


def expanded_dimension_phrase(
    ordered_words: list[dict[str, Any]],
    index: int,
) -> tuple[str, dict[str, float], list[int]]:
    start = index
    end = index

    while start > 0:
        previous = ordered_words[start - 1]
        current = ordered_words[start]
        gap = float(current["x0"]) - float(previous["x1"])
        if gap <= 18 and is_prefix_dimension_token(str(previous.get("text", ""))):
            start -= 1
            continue
        break

    while end + 1 < len(ordered_words):
        current = ordered_words[end]
        next_word = ordered_words[end + 1]
        gap = float(next_word["x0"]) - float(current["x1"])
        next_text = str(next_word.get("text", ""))
        current_text = normalize_text(str(current.get("text", ""))).upper().strip(",")
        next_is_new_dimension, _, _ = looks_like_dimension(next_text)

        include_next = gap <= 30 and (
            is_suffix_dimension_token(next_text)
            or (current_text == "X" and re.match(r"^[0-9]+°?,?$", normalize_text(next_text)))
        )
        if not include_next or (next_is_new_dimension and not is_suffix_dimension_token(next_text)):
            break

        end += 1
        if end - start >= 7:
            break

    phrase_words = ordered_words[start : end + 1]
    phrase = " ".join(str(word.get("text", "")).strip() for word in phrase_words if str(word.get("text", "")).strip())
    return phrase, union_pdf_words(phrase_words), list(range(start, end + 1))


def looks_like_callout_line(line_text: str) -> bool:
    clean = normalize_text(line_text).upper()
    compact = re.sub(r"\s+", "", clean)
    if NOTE_CONTEXT_LINE_RE.search(clean) or GENERAL_TOLERANCE_LINE_RE.search(clean) or HEADER_STATUS_LINE_RE.search(clean):
        return False
    if not re.search(r"(?:^|[\s(])(?:D|R)?\.?\d", clean):
        return False
    return bool(
        re.search(r"\b(?:THRU|ALL|NEAR|FAR|SIDE|TYP|PLCS|PLACES|UNC|UNF|THREAD|THD)\b", clean)
        or re.search(r"\bX\s*\d+(?:\.\d+)?(?:DEG|°)?\b", clean)
        or re.search(r"(?:^|[\s(])[DR]\.?\d", clean)
        or re.search(r"\b\d+\s*-\s*\d+\s*(?:UNC|UNF|UNEF)\b", clean)
        or re.search(r"\.\d{2,4}", compact)
    )


def extract_multiline_callout_candidates(
    lines: list[list[dict[str, Any]]],
    page_index: int,
) -> list[dict[str, Any]]:
    def split_callout_clusters(ordered_line: list[dict[str, Any]]) -> list[list[dict[str, Any]]]:
        clusters: list[list[dict[str, Any]]] = []
        for word in ordered_line:
            if not clusters:
                clusters.append([word])
                continue
            previous = clusters[-1][-1]
            gap = float(word.get("x0", 0)) - float(previous.get("x1", 0))
            if gap > 35:
                clusters.append([word])
            else:
                clusters[-1].append(word)
        return clusters

    candidates: list[dict[str, Any]] = []
    for line_words in lines:
        ordered_line = sorted(line_words, key=lambda item: float(item.get("x0", 0)))
        for cluster_words in split_callout_clusters(ordered_line):
            line_text = " ".join(str(word.get("text", "")).strip() for word in cluster_words if str(word.get("text", "")).strip())
            if not looks_like_callout_line(line_text):
                continue
            ok, confidence, reason = looks_like_dimension(line_text, line_text, line_text)
            if not ok:
                confidence = 0.86
                reason = "technical callout line"
            box = union_pdf_words(cluster_words)
            candidates.append(
                {
                    "page": page_index,
                    "text": line_text,
                    **box,
                    "confidence": round(max(confidence, 0.84), 3),
                    "reason": reason if reason != "not a dimension" else "technical callout line",
                }
            )
    return candidates


def dedupe_candidates(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    def boxes_overlap(a: dict[str, Any], b: dict[str, Any]) -> bool:
        ax0, ay0 = a["x"], a["y"]
        ax1, ay1 = a["x"] + a["width"], a["y"] + a["height"]
        bx0, by0 = b["x"], b["y"]
        bx1, by1 = b["x"] + b["width"], b["y"] + b["height"]
        return ax0 <= bx1 + 1.5 and ax1 + 1.5 >= bx0 and ay0 <= by1 + 1.5 and ay1 + 1.5 >= by0

    def center_distance(a: dict[str, Any], b: dict[str, Any]) -> float:
        ax = a["x"] + a["width"] / 2
        ay = a["y"] + a["height"] / 2
        bx = b["x"] + b["width"] / 2
        by = b["y"] + b["height"] / 2
        return ((ax - bx) ** 2 + (ay - by) ** 2) ** 0.5

    unique: list[dict[str, Any]] = []
    for item in items:
        duplicate = False
        for existing in unique:
            if item["page"] != existing["page"]:
                continue

            same_text = normalize_text(item["text"]).lower() == normalize_text(existing["text"]).lower()
            same_physical_text = boxes_overlap(item, existing) or center_distance(item, existing) <= 4
            if same_text and same_physical_text:
                duplicate = True
                break
            item_text = re.sub(r"\s+", "", normalize_text(item["text"]).lower())
            existing_text = re.sub(r"\s+", "", normalize_text(existing["text"]).lower())
            text_contained = item_text in existing_text or existing_text in item_text
            if same_physical_text and text_contained:
                if len(item_text) > len(existing_text):
                    unique.remove(existing)
                    break
                duplicate = True
                break

        if not duplicate:
            unique.append(item)

    return unique


def remove_grouped_digit_artifacts(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cleaned: list[dict[str, Any]] = []
    for item in items:
        compact = re.sub(r"\s+", "", normalize_text(str(item.get("text", ""))))
        if item.get("reason") == "grouped digit fragments" and re.match(r"^[0-9]+$", compact):
            continue
        cleaned.append(item)
    return cleaned


def filter_candidates_for_profile(items: list[dict[str, Any]], profile: AnalysisProfile) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for item in items:
        confidence = float(item.get("confidence", 0))
        if confidence < profile.min_confidence:
            continue
        reason = str(item.get("reason", ""))
        text = normalize_text(str(item.get("text", ""))).strip()
        compact = re.sub(r"\s+", "", text)
        is_plain_number = bool(re.match(r"^[0-9]+(?:[.,][0-9]+)?$", compact))
        if not profile.allow_plain_numbers and is_plain_number and "unit" not in reason:
            continue
        filtered.append(item)
    return filtered


def finalize_raw_candidates(
    raw_candidates: list[dict[str, Any]],
    profile: AnalysisProfile,
) -> list[DimensionCandidate]:
    raw_candidates = merge_digit_fragments(raw_candidates)
    raw_candidates = remove_grouped_digit_artifacts(raw_candidates)
    raw_candidates = [item for item in raw_candidates if not is_date_like_text(str(item.get("text", "")))]
    raw_candidates = filter_candidates_for_profile(raw_candidates, profile)
    raw_candidates = dedupe_candidates(raw_candidates)
    raw_candidates = visual_order_candidates(raw_candidates)
    return [
        DimensionCandidate(number=index, **item)
        for index, item in enumerate(raw_candidates, start=1)
    ]


def score_candidate_set(candidates: list[DimensionCandidate], profile: AnalysisProfile) -> float:
    if not candidates:
        return 0.0

    count = len(candidates)
    average_confidence = sum(candidate.confidence for candidate in candidates) / count
    technical_markers = sum(
        1
        for candidate in candidates
        if re.search(r"(?i)(^\(?[rdm]|x|\+/-|\u00b1|\+|-|/|mm|cm|in|deg|grados|\")", normalize_text(candidate.text))
    )
    marker_ratio = technical_markers / count
    count_score = min(count, 80) / 80
    profile_bias = 0.04 if profile.name == "standard" else 0.0
    if profile.name == "permissive" and count > 140:
        profile_bias -= 0.18
    if profile.name == "conservative" and count < 8:
        profile_bias -= 0.08
    return (average_confidence * 0.56) + (marker_ratio * 0.28) + (count_score * 0.16) + profile_bias


def resolve_analysis_profile(strategy: str) -> AnalysisProfile:
    key = strategy.strip().lower()
    if key not in ANALYSIS_PROFILES:
        raise ValueError("Estrategia no valida. Use auto, standard, conservative, permissive, vertical_side u ocr.")
    return ANALYSIS_PROFILES[key]


def visual_order_candidates(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ordered: list[dict[str, Any]] = []
    by_page: dict[int, list[dict[str, Any]]] = {}
    for item in items:
        by_page.setdefault(int(item["page"]), []).append(item)

    for page in sorted(by_page):
        page_items = sorted(
            by_page[page],
            key=lambda item: (
                float(item["y"]) + float(item.get("height", 0)) / 2,
                float(item["x"]) + float(item.get("width", 0)) / 2,
            ),
        )
        heights = sorted(float(item.get("height", 0)) for item in page_items if float(item.get("height", 0)) > 0)
        median_height = heights[len(heights) // 2] if heights else 8.0
        # Number plans the way a machinist scans them: top-to-bottom bands,
        # then left-to-right inside each band. A wide band made lower left
        # dimensions jump ahead of upper dimensions, so keep it tight.
        band_tolerance = min(48.0, max(28.0, median_height * 4.5))
        bands: list[dict[str, Any]] = []
        for item in page_items:
            center_y = float(item["y"]) + float(item.get("height", 0)) / 2
            if not bands:
                bands.append({"items": [item], "min_y": center_y, "max_y": center_y})
                continue

            current_band = bands[-1]
            if center_y - float(current_band["min_y"]) <= band_tolerance:
                current_band["items"].append(item)
                current_band["min_y"] = min(float(current_band["min_y"]), center_y)
                current_band["max_y"] = max(float(current_band["max_y"]), center_y)
            else:
                bands.append({"items": [item], "min_y": center_y, "max_y": center_y})

        for band in sorted(bands, key=lambda entry: float(entry["min_y"])):
            band_items = list(band["items"])
            ordered.extend(
                sorted(
                    band_items,
                    key=lambda item: (
                        float(item["x"]) + float(item.get("width", 0)) / 2,
                        float(item["y"]) + float(item.get("height", 0)) / 2,
                    ),
                )
            )

    return ordered


def union_word_box(words: list[dict[str, Any]]) -> dict[str, float]:
    xs = [float(word["x"]) for word in words]
    ys = [float(word["y"]) for word in words]
    rights = [float(word["x"]) + float(word["width"]) for word in words]
    bottoms = [float(word["y"]) + float(word["height"]) for word in words]
    return {
        "x": min(xs),
        "y": min(ys),
        "width": max(rights) - min(xs),
        "height": max(bottoms) - min(ys),
    }


def merge_ocr_decimal_words(words: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    by_line: dict[str, list[dict[str, Any]]] = {}
    for word in words:
        by_line.setdefault(str(word.get("line", "")), []).append(word)

    for line, line_words in by_line.items():
        ordered = sorted(line_words, key=lambda item: float(item["x"]))
        used: set[int] = set()
        for index, word in enumerate(ordered):
            if index in used:
                continue

            text = normalize_text(str(word.get("text", "")))
            if (
                index + 1 < len(ordered)
                and re.match(r"^[0-9]{2,3}$", text)
                and re.match(r"^[0-9]$", normalize_text(str(ordered[index + 1].get("text", ""))))
            ):
                next_word = ordered[index + 1]
                gap = float(next_word["x"]) - (float(word["x"]) + float(word["width"]))
                same_baseline = abs(float(next_word["y"]) - float(word["y"])) <= 10
                if 0 <= gap <= 24 and same_baseline:
                    box = union_word_box([word, next_word])
                    merged.append(
                        {
                            **word,
                            **box,
                            "text": f"{text}.{normalize_text(str(next_word.get('text', '')))}",
                            "line": line,
                        }
                    )
                    used.update({index, index + 1})
                    continue

            merged.append(word)
            used.add(index)

    return merged


def merge_digit_fragments(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    singles = [
        item
        for item in items
        if re.match(r"^[0-9]$", normalize_text(item["text"])) and item.get("confidence", 0) <= 0.75
    ]
    consumed: set[int] = set()
    merged: list[dict[str, Any]] = []

    def center(item: dict[str, Any]) -> tuple[float, float]:
        return (item["x"] + item["width"] / 2, item["y"] + item["height"] / 2)

    for index, item in enumerate(items):
        if index in consumed:
            continue

        if item not in singles:
            merged.append(item)
            continue

        x, y = center(item)
        group_indexes = [index]
        for other_index, other in enumerate(items):
            if other_index == index or other_index in consumed or other not in singles:
                continue
            if other["page"] != item["page"]:
                continue

            ox, oy = center(other)
            same_vertical_number = abs(ox - x) <= 8 and abs(oy - y) <= 42
            same_horizontal_number = abs(oy - y) <= 8 and abs(ox - x) <= 42
            if same_vertical_number or same_horizontal_number:
                group_indexes.append(other_index)

        if len(group_indexes) < 2:
            merged.append(item)
            continue

        group = [items[group_index] for group_index in group_indexes]
        xs = [entry["x"] for entry in group]
        ys = [entry["y"] for entry in group]
        rights = [entry["x"] + entry["width"] for entry in group]
        bottoms = [entry["y"] + entry["height"] for entry in group]
        vertical = (max(xs) - min(xs)) <= 8
        ordered = sorted(group, key=lambda entry: entry["y"] if vertical else entry["x"])

        for group_index in group_indexes:
            consumed.add(group_index)

        merged.append(
            {
                "page": item["page"],
                "text": "".join(entry["text"] for entry in ordered),
                "x": min(xs),
                "y": min(ys),
                "width": max(rights) - min(xs),
                "height": max(bottoms) - min(ys),
                "confidence": max(entry["confidence"] for entry in group),
                "reason": "grouped digit fragments",
            }
        )

    return merged


def preprocess_for_ocr(source: Path, target: Path, scale: int = 5) -> None:
    image = Image.open(source).convert("L")
    image = ImageOps.autocontrast(image)
    image = ImageEnhance.Contrast(image).enhance(2.2)
    image = image.resize((image.width * scale, image.height * scale), Image.Resampling.LANCZOS)
    image = image.filter(ImageFilter.SHARPEN)
    
    target.parent.mkdir(parents=True, exist_ok=True)
    image.save(target)


def build_ocr_variants(source: Path, output_dir: Path) -> list[tuple[Path, tuple[float, float, float, float], int]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    variants: list[tuple[Path, tuple[float, float, float, float], int]] = []

    with Image.open(source) as original:
        width, height = original.size
        crops = [(0, 0, width, height)]

        # Overlapping zones help Windows OCR detect small drawing dimensions.
        half_w = width // 2
        half_h = height // 2
        overlap_x = max(40, width // 12)
        overlap_y = max(30, height // 12)
        crops.extend(
            [
                (0, 0, min(width, half_w + overlap_x), min(height, half_h + overlap_y)),
                (max(0, half_w - overlap_x), 0, width, min(height, half_h + overlap_y)),
                (0, max(0, half_h - overlap_y), min(width, half_w + overlap_x), height),
                (max(0, half_w - overlap_x), max(0, half_h - overlap_y), width, height),
            ]
        )

        if width > 700 or height > 400:
            third_w = width // 3
            third_h = height // 3
            for row in range(3):
                for col in range(3):
                    left = max(0, col * third_w - overlap_x)
                    top = max(0, row * third_h - overlap_y)
                    right = min(width, (col + 1) * third_w + overlap_x)
                    bottom = min(height, (row + 1) * third_h + overlap_y)
                    crops.append((left, top, right, bottom))

        seen: set[tuple[int, int, int, int]] = set()
        for index, crop_box in enumerate(crops):
            normalized = tuple(int(value) for value in crop_box)
            if normalized in seen:
                continue
            seen.add(normalized)

            crop = original.crop(normalized)
            rotations = (0, 90, 270) if index <= 4 else (0,)
            for rotation in rotations:
                rotated = crop if rotation == 0 else crop.rotate(rotation, expand=True)
                crop_path = output_dir / f"{source.stem}_crop{index}_rot{rotation}.png"
                raw_crop_path = output_dir / f"{source.stem}_crop{index}_rot{rotation}_raw.png"
                rotated.save(raw_crop_path)
                preprocess_for_ocr(raw_crop_path, crop_path)
                raw_crop_path.unlink(missing_ok=True)
                variants.append((crop_path, normalized, rotation))

    return variants


def run_windows_ocr(image_path: Path) -> list[dict[str, Any]]:
    script = Path(__file__).with_name("windows_ocr.ps1")
    powershell = shutil.which("powershell") or shutil.which("pwsh")
    if not script.exists() or not powershell:
        return []

    try:
        result = subprocess.run(
            [
                powershell,
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script),
                str(image_path),
            ],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=60,
        )
    except (FileNotFoundError, subprocess.SubprocessError):
        return []
    if result.returncode != 0 or not result.stdout.strip():
        return []

    payload = json.loads(result.stdout, strict=False)
    if isinstance(payload, dict):
        return [payload]
    return payload


def extracted_page_images(pdf_path: Path, output_dir: Path) -> dict[int, list[Path]]:
    reader = PdfReader(str(pdf_path))
    extracted: dict[int, list[Path]] = {}
    output_dir.mkdir(parents=True, exist_ok=True)

    for page_index, page in enumerate(reader.pages, start=1):
        page_images: list[Path] = []
        for image_index, image in enumerate(page.images, start=1):
            ext = Path(image.name).suffix or ".jpg"
            target = output_dir / f"page{page_index}_image{image_index}{ext}"
            target.write_bytes(image.data)
            page_images.append(target)
        extracted[page_index] = page_images

    return extracted


def extract_ocr_candidates(
    pdf_path: Path,
    profile: AnalysisProfile | None = None,
) -> list[DimensionCandidate]:
    profile = profile or ANALYSIS_PROFILES["ocr"]
    raw_candidates: list[dict[str, Any]] = []

    with tempfile.TemporaryDirectory(prefix="cotas_ocr_") as temp_name:
        temp_dir = Path(temp_name)
        images_by_page = extracted_page_images(pdf_path, temp_dir / "images")

        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                page_images = images_by_page.get(page_index, [])
                for image_number, page_image in enumerate(page_images):
                    if image_number >= len(page.images):
                        continue

                    image_meta = page.images[image_number]
                    pdf_x0 = float(image_meta["x0"])
                    pdf_top = float(image_meta["top"])
                    pdf_width = float(image_meta["width"])
                    pdf_height = float(image_meta["height"])
                    with Image.open(page_image) as original_image:
                        original_width, original_height = original_image.size

                    variants = build_ocr_variants(page_image, temp_dir / "ocr" / page_image.stem)
                    for ocr_image, crop_box, rotation in variants:
                        ocr_words = merge_ocr_decimal_words(run_windows_ocr(ocr_image))
                        if not ocr_words:
                            continue

                        crop_left, crop_top, crop_right, crop_bottom = crop_box
                        with Image.open(ocr_image) as processed:
                            image_width, image_height = processed.size
                        crop_width = crop_right - crop_left
                        crop_height = crop_bottom - crop_top

                        for word in ocr_words:
                            text = str(word.get("text", "")).strip()
                            line_text = str(word.get("line", ""))
                            ok, confidence, reason = looks_like_dimension(text, line_text, line_text)
                            if not ok:
                                continue

                            word_x = float(word["x"]) / image_width
                            word_y = float(word["y"]) / image_height
                            word_w = float(word["width"]) / image_width
                            word_h = float(word["height"]) / image_height

                            if rotation == 0:
                                original_x = crop_left + word_x * crop_width
                                original_y = crop_top + word_y * crop_height
                                original_w = word_w * crop_width
                                original_h = word_h * crop_height
                            elif rotation == 90:
                               original_x = crop_left + (1 - word_y - word_h) * crop_width
                               original_y = crop_top + word_x * crop_height
                               original_w = word_h * crop_width
                               original_h = word_w * crop_height
                            else:  # 270
                               original_x = crop_left + word_y * crop_width
                               original_y = crop_top + (1 - word_x - word_w) * crop_height
                               original_w = word_h * crop_width
                               original_h = word_w * crop_height

                            raw_candidates.append(
                                {
                                    "page": page_index,
                                    "text": text,
                                    "x": pdf_x0 + (original_x / original_width) * pdf_width,
                                    "y": pdf_top + (original_y / original_height) * pdf_height,
                                    "width": (original_w / original_width) * pdf_width,
                                    "height": (original_h / original_height) * pdf_height,
                                    "confidence": round(max(confidence - 0.1, 0.5), 3),
                                    "reason": f"ocr {reason}",
                                }
                            )

    return finalize_raw_candidates(raw_candidates, profile)


def extract_candidates_with_profile(pdf_path: Path, profile: AnalysisProfile) -> list[DimensionCandidate]:
    if profile.use_ocr:
        return extract_ocr_candidates(pdf_path, profile)

    raw_candidates: list[dict[str, Any]] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_index, page in enumerate(pdf.pages, start=1):
            table_bboxes = [] if profile.include_tables else detected_table_bboxes(page)
            all_words = page.extract_words(
                keep_blank_chars=False,
                use_text_flow=False,
                extra_attrs=[],
            )
            general_tolerance_bboxes = [] if profile.include_tables else detected_general_tolerance_bboxes(
                all_words,
                float(page.width),
                float(page.height),
            )
            title_block_bboxes = detected_title_block_bboxes(
                all_words,
                float(page.width),
                float(page.height),
            ) if profile.exclude_title_block else []
            note_bboxes = detected_note_bboxes(
                all_words,
                float(page.width),
                float(page.height),
            ) if profile.exclude_title_block else []
            profile_bboxes = profile_excluded_bboxes(profile, float(page.width), float(page.height))
            excluded_bboxes = [
                *table_bboxes,
                *general_tolerance_bboxes,
                *title_block_bboxes,
                *note_bboxes,
                *profile_bboxes,
            ]
            if profile.use_rotated:
                raw_candidates.extend(
                    extract_rotated_word_candidates(
                        all_words,
                        page_index,
                        float(page.width),
                        float(page.height),
                        excluded_bboxes,
                    )
                )
            if profile.use_stacked:
                raw_candidates.extend(
                    extract_stacked_autocad_decimal_candidates(
                        page.chars,
                        page_index,
                        excluded_bboxes,
                    )
                )
            words = [
                word
                for word in all_words
                if bool(word.get("upright", True)) and not word_inside_any_bbox(word, excluded_bboxes)
            ]
            lines: list[list[dict[str, Any]]] = []
            for word in sorted(words, key=lambda item: (float(item.get("top", 0)), float(item.get("x0", 0)))):
                if not lines or abs(float(lines[-1][0].get("top", 0)) - float(word.get("top", 0))) >= 3:
                    lines.append([word])
                else:
                    lines[-1].append(word)

            raw_candidates.extend(extract_multiline_callout_candidates(lines, page_index))

            for line_words in lines:
                ordered_line = sorted(line_words, key=lambda item: float(item.get("x0", 0)))
                used_indexes: set[int] = set()
                line_text = " ".join(str(peer.get("text", "")) for peer in ordered_line)
                for index, word in enumerate(ordered_line):
                    if index in used_indexes:
                        continue

                    text = str(word.get("text", "")).strip()
                    compact_check = re.sub(r"\s+", "", text)
                    if re.match(r"^[0-9]{1,2}$", compact_check):
                        margin = page.height * 0.07
                        if word["top"] <= margin or word["bottom"] >= page.height - margin:
                            continue
                    neighbor_text = nearby_context(ordered_line, word)
                    ok, confidence, reason = looks_like_dimension(text, line_text, neighbor_text)
                    if not ok:
                        continue

                    phrase, box, phrase_indexes = expanded_dimension_phrase(ordered_line, index)
                    if phrase and phrase != text:
                        reason = "dimension phrase"
                        confidence = min(confidence + 0.08, 0.96)
                        used_indexes.update(phrase_indexes)

                    in_title_block = box["x"] >= float(page.width) * 0.55 and box["y"] >= float(page.height) * 0.86
                    if profile.exclude_title_block and in_title_block:
                        continue

                    compact_phrase = re.sub(r"\s+", "", normalize_text(phrase or text))
                    is_stacked_tolerance = bool(re.match(r"^\.(?:0{2,3}|00[1-9])$", compact_phrase)) and any(
                        str(peer.get("text", "")).strip() in {"+", "-"}
                        and abs(float(peer.get("x0", 0)) - box["x"]) <= 28
                        and abs(float(peer.get("top", 0)) - box["y"]) <= 18
                        for peer in words
                    )
                    if is_stacked_tolerance:
                        continue

                    raw_candidates.append(
                        {
                            "page": page_index,
                            "text": phrase or text,
                            **box,
                            "confidence": round(confidence, 3),
                            "reason": reason,
                        }
                    )

    candidates = finalize_raw_candidates(raw_candidates, profile)
    if not candidates and profile.name != "ocr":
        return extract_ocr_candidates(pdf_path)

    return candidates


def extract_candidates(
    pdf_path: Path,
    include_tables: bool = False,
    strategy: str = DEFAULT_ANALYSIS_STRATEGY,
) -> tuple[list[DimensionCandidate], str, bool]:
    if include_tables and strategy == DEFAULT_ANALYSIS_STRATEGY:
        strategy = "permissive"

    if strategy != DEFAULT_ANALYSIS_STRATEGY:
        profile = resolve_analysis_profile(strategy)
        candidates = extract_candidates_with_profile(pdf_path, profile)
        return candidates, profile.name, profile.use_ocr or any("ocr" in candidate.reason for candidate in candidates)

    scored_results: list[tuple[float, AnalysisProfile, list[DimensionCandidate]]] = []
    for profile_name in ("standard", "conservative", "permissive"):
        profile = ANALYSIS_PROFILES[profile_name]
        candidates = extract_candidates_with_profile(pdf_path, profile)
        scored_results.append((score_candidate_set(candidates, profile), profile, candidates))

    best_score, best_profile, best_candidates = max(scored_results, key=lambda item: item[0])
    if not best_candidates or best_score < 0.46:
        ocr_profile = ANALYSIS_PROFILES["ocr"]
        ocr_candidates = extract_ocr_candidates(pdf_path, ocr_profile)
        ocr_score = score_candidate_set(ocr_candidates, ocr_profile)
        if ocr_score >= best_score or not best_candidates:
            return ocr_candidates, ocr_profile.name, True

    return best_candidates, best_profile.name, any("ocr" in candidate.reason for candidate in best_candidates)


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(value, maximum))


def label_font_size_for_page(page_width: float, page_height: float, number: int) -> float:
    base_size = 8.6 if number < 100 else 7.8
    sheet_scale = clamp(min(page_width, page_height) / 800.0, 1.0, 1.25)
    return round(base_size * sheet_scale, 2)


def boxes_intersect(a: PdfBBox, b: PdfBBox, padding: float = 0.0) -> bool:
    ax0, at, ax1, ab = a
    bx0, bt, bx1, bb = b
    return (
        ax0 - padding <= bx1
        and ax1 + padding >= bx0
        and at - padding <= bb
        and ab + padding >= bt
    )


def box_area(box: PdfBBox) -> float:
    x0, top, x1, bottom = box
    return max(0.0, x1 - x0) * max(0.0, bottom - top)


def intersection_area(a: PdfBBox, b: PdfBBox, padding: float = 0.0) -> float:
    ax0, at, ax1, ab = a
    bx0, bt, bx1, bb = b
    left = max(ax0 - padding, bx0)
    top = max(at - padding, bt)
    right = min(ax1 + padding, bx1)
    bottom = min(ab + padding, bb)
    return max(0.0, right - left) * max(0.0, bottom - top)


def point_in_box(x: float, y: float, box: PdfBBox, padding: float = 0.0) -> bool:
    x0, top, x1, bottom = box
    return (x0 - padding) <= x <= (x1 + padding) and (top - padding) <= y <= (bottom + padding)


def segment_intersects_box(
    start: tuple[float, float],
    end: tuple[float, float],
    box: PdfBBox,
    padding: float = 0.0,
) -> bool:
    x0, y0 = start
    x1, y1 = end
    bx0, bt, bx1, bb = box
    bx0 -= padding
    bt -= padding
    bx1 += padding
    bb += padding

    if point_in_box(x0, y0, (bx0, bt, bx1, bb)) or point_in_box(x1, y1, (bx0, bt, bx1, bb)):
        return True

    dx = x1 - x0
    dy = y1 - y0
    checks: list[tuple[str, float]] = []
    if dx:
        checks.extend([("x", bx0), ("x", bx1)])
    if dy:
        checks.extend([("y", bt), ("y", bb)])

    for axis, value in checks:
        t = (value - x0) / dx if axis == "x" else (value - y0) / dy
        if 0.0 <= t <= 1.0:
            x = x0 + t * dx
            y = y0 + t * dy
            if bx0 <= x <= bx1 and bt <= y <= bb:
                return True
    return False


def line_start_outside_source_box(
    source_box: PdfBBox,
    label_point: tuple[float, float],
    page_height: float,
    padding: float = 3.0,
) -> tuple[float, float]:
    x0, top, x1, bottom = source_box
    center_x = (x0 + x1) / 2
    center_top_y = (top + bottom) / 2
    label_x, label_pdf_y = label_point
    label_top_y = page_height - label_pdf_y

    dx = label_x - center_x
    dy = label_top_y - center_top_y
    if abs(dx) >= abs(dy):
        start_x = x1 + padding if dx >= 0 else x0 - padding
        scale = (start_x - center_x) / dx if dx else 0
        start_top_y = center_top_y + dy * scale
        start_top_y = clamp(start_top_y, top - padding, bottom + padding)
    else:
        start_top_y = bottom + padding if dy >= 0 else top - padding
        scale = (start_top_y - center_top_y) / dy if dy else 0
        start_x = center_x + dx * scale
        start_x = clamp(start_x, x0 - padding, x1 + padding)

    return start_x, page_height - start_top_y


def page_text_bboxes(pdf_path: Path) -> dict[int, list[PdfBBox]]:
    occupied: dict[int, list[PdfBBox]] = {}
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                boxes = []
                for word in page.extract_words(keep_blank_chars=False, use_text_flow=True, extra_attrs=[]):
                    boxes.append(
                        (
                            float(word["x0"]),
                            float(word["top"]),
                            float(word["x1"]),
                            float(word["bottom"]),
                        )
                    )
                occupied[page_index] = boxes
    except Exception:
        return {}
    return occupied


def page_graphic_bboxes(pdf_path: Path) -> dict[int, list[PdfBBox]]:
    occupied: dict[int, list[PdfBBox]] = {}
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                boxes: list[PdfBBox] = []
                for item in [*page.lines, *page.curves, *page.rects]:
                    x0 = float(item.get("x0", 0))
                    x1 = float(item.get("x1", x0))
                    top = float(item.get("top", 0))
                    bottom = float(item.get("bottom", top))
                    if x1 < x0:
                        x0, x1 = x1, x0
                    if bottom < top:
                        top, bottom = bottom, top
                    width = x1 - x0
                    height = bottom - top
                    if width <= 0.1 and height <= 0.1:
                        continue

                    boxes.append((x0, top, x1, bottom))
                occupied[page_index] = boxes
    except Exception:
        return {}
    return occupied


def page_forbidden_label_bboxes(pdf_path: Path) -> dict[int, list[PdfBBox]]:
    occupied: dict[int, list[PdfBBox]] = {}
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                words = page.extract_words(
                    keep_blank_chars=False,
                    use_text_flow=False,
                    extra_attrs=[],
                )
                occupied[page_index] = [
                    *detected_title_block_bboxes(words, float(page.width), float(page.height)),
                    *detected_note_bboxes(words, float(page.width), float(page.height)),
                    *detected_general_tolerance_bboxes(words, float(page.width), float(page.height)),
                    *detected_revision_history_bboxes(words, float(page.width), float(page.height)),
                ]
    except Exception:
        return {}
    return occupied


def expanded_horizontal_source_box(candidate: DimensionCandidate, text_boxes: list[PdfBBox]) -> PdfBBox:
    box = (
        candidate.x,
        candidate.y,
        candidate.x + candidate.width,
        candidate.y + candidate.height,
    )
    if "rotated" in candidate.reason:
        return box

    text = normalize_text(candidate.text).upper()
    expands_as_multiline_callout = "THRU" in text or "90" in text
    if not expands_as_multiline_callout:
        return box

    center_top_y = candidate.y + candidate.height / 2
    same_line_boxes: list[PdfBBox] = []
    for text_box in text_boxes:
        x0, top, x1, bottom = text_box
        text_center_y = (top + bottom) / 2
        same_row = abs(text_center_y - center_top_y) <= max(6.0, candidate.height * 0.7)
        close_x = x1 >= candidate.x - 36 and x0 <= candidate.x + candidate.width + 14
        if same_row and close_x:
            same_line_boxes.append(text_box)

    left_symbol_guard = 16.0

    if not same_line_boxes:
        return (
            max(0.0, box[0] - left_symbol_guard),
            box[1],
            box[2],
            box[3],
        )

    return (
        max(0.0, min(candidate.x, *(item[0] for item in same_line_boxes)) - left_symbol_guard),
        min(candidate.y, *(item[1] for item in same_line_boxes)),
        max(candidate.x + candidate.width, *(item[2] for item in same_line_boxes)),
        max(candidate.y + candidate.height, *(item[3] for item in same_line_boxes)),
    )


def choose_label_position(
    candidate: DimensionCandidate,
    page_width: float,
    page_height: float,
    label_half_width: float,
    label_half_height: float,
    text_boxes: list[PdfBBox],
    graphic_boxes: list[PdfBBox],
    placed_label_boxes: list[PdfBBox],
    forbidden_label_boxes: list[PdfBBox] | None = None,
) -> tuple[float, float]:
    forbidden_label_boxes = forbidden_label_boxes or []
    center_x = candidate.x + candidate.width / 2
    center_top_y = candidate.y + candidate.height / 2
    expanded_box = expanded_horizontal_source_box(candidate, text_boxes)
    source_box = (
        expanded_box[0] - 3,
        expanded_box[1] - 3,
        expanded_box[2] + 3,
        expanded_box[3] + 3,
    )
    rings = [9, 12, 16, 22, 30, 40]
    directions = [
        (1.0, -0.8),
        (-1.0, -0.8),
        (1.0, 0.8),
        (-1.0, 0.8),
        (1.0, 0.0),
        (-1.0, 0.0),
        (0.0, -1.0),
        (0.0, 1.0),
        (1.5, -0.5),
        (-1.5, -0.5),
        (1.5, 0.5),
        (-1.5, 0.5),
        (0.5, -1.5),
        (-0.5, -1.5),
        (0.5, 1.5),
        (-0.5, 1.5),
    ]

    best_x = clamp(center_x + rings[0], label_half_width + 4, page_width - label_half_width - 4)
    best_top_y = clamp(center_top_y - rings[0], label_half_height + 4, page_height - label_half_height - 4)
    best_score = float("inf")
    candidate_points: list[tuple[float, float, float, str]] = []

    # Shop-friendly default: place numbers on the same horizontal line as the
    # dimension value, either to its left or right.
    is_rotated_dimension = "rotated" in candidate.reason
    if is_rotated_dimension:
        for gap in [7, 10, 14, 20, 28, 38]:
            candidate_points.append((center_x, candidate.y - gap, gap, "rotated_axis_above"))
            candidate_points.append((center_x, candidate.y + candidate.height + gap, gap, "rotated_axis_below"))

    min_horizontal_gap = 18 if is_rotated_dimension else max(12.0, label_half_width + 6.0)
    for gap in [8, 11, 15, 21, 30, 40]:
        if gap < min_horizontal_gap:
            continue
        candidate_points.append((expanded_box[0] - gap, center_top_y, gap, "horizontal"))
        candidate_points.append((expanded_box[2] + gap, center_top_y, gap, "horizontal"))

    label_text_boxes = text_boxes
    leader_text_boxes = [
        box
        for box in text_boxes
        if not boxes_intersect(box, source_box, padding=2.0)
    ]

    for ring in rings:
        for dir_x, dir_y in directions:
            length = (dir_x * dir_x + dir_y * dir_y) ** 0.5 or 1.0
            dx = ring * dir_x / length
            dy = ring * dir_y / length
            candidate_points.append((center_x + dx, center_top_y + dy, abs(dx) + abs(dy), "local"))

    fallback_x = best_x
    fallback_top_y = best_top_y
    fallback_score = float("inf")
    for raw_x, raw_top_y, distance, placement_mode in candidate_points:
            x = clamp(raw_x, label_half_width + 4, page_width - label_half_width - 4)
            top_y = clamp(raw_top_y, label_half_height + 4, page_height - label_half_height - 4)
            label_box = (
                x - label_half_width - 2,
                top_y - label_half_height - 2,
                x + label_half_width + 2,
                top_y + label_half_height + 2,
            )

            text_overlap = sum(intersection_area(label_box, box, padding=3.0) for box in label_text_boxes)
            label_overlap = sum(intersection_area(label_box, box, padding=4.0) for box in placed_label_boxes)
            source_overlap = intersection_area(label_box, source_box, padding=2.0)
            text_hits = sum(1 for box in label_text_boxes if boxes_intersect(label_box, box, padding=3.0))
            label_hits = sum(1 for box in placed_label_boxes if boxes_intersect(label_box, box, padding=4.0))
            forbidden_hits = sum(
                1 for box in forbidden_label_boxes if boxes_intersect(label_box, box, padding=4.0)
            )
            source_hit = boxes_intersect(label_box, source_box, padding=2.0)
            graphic_hits = sum(1 for box in graphic_boxes if boxes_intersect(label_box, box, padding=2.0))

            source_pdf_box = (
                expanded_box[0],
                expanded_box[1],
                expanded_box[2],
                expanded_box[3],
            )
            line_start_x, line_start_pdf_y = line_start_outside_source_box(
                source_pdf_box,
                (x, page_height - top_y),
                page_height,
            )
            leader_start = (line_start_x, page_height - line_start_pdf_y)
            leader_end = (x, top_y)
            leader_length = ((leader_end[0] - leader_start[0]) ** 2 + (leader_end[1] - leader_start[1]) ** 2) ** 0.5
            max_leader = 56 if is_rotated_dimension else 48
            if forbidden_hits:
                continue
            fallback_score_candidate = leader_length * 1000 + text_hits * 9000 + label_hits * 15000 + source_overlap * 200
            if fallback_score_candidate < fallback_score:
                fallback_score = fallback_score_candidate
                fallback_x = x
                fallback_top_y = top_y
            if leader_length > max_leader:
                continue
            leader_text_hits = sum(
                1
                for box in leader_text_boxes
                if segment_intersects_box(leader_start, leader_end, box, padding=1.2)
            )
            leader_label_hits = sum(
                1
                for box in placed_label_boxes
                if segment_intersects_box(leader_start, leader_end, box, padding=2.0)
            )
            leader_graphic_hits = sum(
                1
                for box in graphic_boxes
                if segment_intersects_box(leader_start, leader_end, box, padding=1.2)
            )

            edge_bonus = 0
            if x < page_width * 0.12 or x > page_width * 0.88 or top_y < page_height * 0.12 or top_y > page_height * 0.88:
                edge_bonus = -120
            min_leader = 12 if is_rotated_dimension else 9
            short_leader_penalty = (min_leader - leader_length) * 350 if leader_length < min_leader else 0
            soft_max_leader = 34 if placement_mode == "horizontal" else 42
            hard_max_leader = 58 if placement_mode == "horizontal" else 68
            long_leader_penalty = 0
            if leader_length > soft_max_leader:
                long_leader_penalty += (leader_length - soft_max_leader) * 2200
            if leader_length > hard_max_leader:
                long_leader_penalty += 300000 + (leader_length - hard_max_leader) * 5000
            horizontal = placement_mode == "horizontal" and abs(top_y - center_top_y) <= 2.5
            clean_horizontal = (
                horizontal
                and not source_hit
                and text_hits == 0
                and label_hits == 0
                and leader_text_hits == 0
                and leader_label_hits == 0
                and graphic_hits <= 1
                and leader_length >= min_leader
                and leader_length <= hard_max_leader
            )
            horizontal_bonus = 0
            if clean_horizontal:
                horizontal_bonus = -28000
                if is_rotated_dimension:
                    horizontal_bonus -= 2500
            elif horizontal and not source_hit and text_hits == 0 and label_hits == 0 and leader_text_hits == 0:
                horizontal_bonus = -12000
            axis_aligned = placement_mode.startswith("rotated_axis") and abs(x - center_x) <= 2.5
            if axis_aligned and text_hits == 0 and label_hits == 0 and leader_text_hits == 0 and leader_label_hits == 0:
                horizontal_bonus -= 24000
            elif axis_aligned and text_hits == 0 and label_hits == 0:
                horizontal_bonus -= 7000
            if placement_mode == "rotated_axis_above" and text_hits == 0 and label_hits == 0:
                horizontal_bonus -= 30000

            distance_weight = 950 if placement_mode == "horizontal" else 900

            score = (
                text_hits * 25000
                + label_hits * 40000
                + forbidden_hits * 10000000
                + (100000 if source_hit else 0)
                + graphic_hits * 6500
                + text_overlap * 200
                + label_overlap * 300
                + source_overlap * 500
                + leader_text_hits * 4500
                + leader_label_hits * 4500
                + leader_graphic_hits * 1200
                + distance * distance_weight
                + leader_length * 1200
                + edge_bonus
                + horizontal_bonus
                + short_leader_penalty
                + long_leader_penalty
            )
            if score < best_score:
                best_score = score
                best_x = x
                best_top_y = top_y

    if best_score == float("inf"):
        best_x = fallback_x
        best_top_y = fallback_top_y

    return best_x, page_height - best_top_y


def draw_numbered_overlay(
    original_pdf: Path,
    candidates: list[DimensionCandidate],
    output_pdf: Path,
) -> None:
    reader = PdfReader(str(original_pdf))
    writer = PdfWriter()
    occupied_by_page = page_text_bboxes(original_pdf)
    graphics_by_page = page_graphic_bboxes(original_pdf)
    forbidden_labels_by_page = page_forbidden_label_bboxes(original_pdf)
    by_page: dict[int, list[DimensionCandidate]] = {}
    for candidate in candidates:
        by_page.setdefault(candidate.page, []).append(candidate)

    tmp_dir = output_pdf.parent / "_tmp_overlay"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    for page_index, page in enumerate(reader.pages, start=1):
        if int(page.get("/Rotate", 0) or 0) % 360:
            page.transfer_rotation_to_content()
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        overlay_path = tmp_dir / f"overlay_{page_index}.pdf"

        c = canvas.Canvas(str(overlay_path), pagesize=(width, height))
        c.setLineWidth(0.55)
        forbidden_label_boxes = list(forbidden_labels_by_page.get(page_index, []))
        placed_label_boxes: list[PdfBBox] = list(forbidden_label_boxes)

        for candidate in by_page.get(page_index, []):
            font_size = label_font_size_for_page(width, height, candidate.number)
            label = str(candidate.number)
            c.setFont("Helvetica-Bold", font_size)
            text_width = c.stringWidth(label, "Helvetica-Bold", font_size)
            label_half_width = text_width / 2 + 1.5
            label_half_height = font_size / 2 + 1.5
            text_x = candidate.x + candidate.width / 2
            text_y = height - candidate.y - (candidate.height / 2)
            pdf_x, pdf_y = choose_label_position(
                candidate,
                width,
                height,
                label_half_width,
                label_half_height,
                occupied_by_page.get(page_index, []),
                graphics_by_page.get(page_index, []),
                placed_label_boxes,
                forbidden_label_boxes,
            )
            label_top_y = height - pdf_y
            placed_label_boxes.append(
                (
                    pdf_x - label_half_width - 2,
                    label_top_y - label_half_height - 2,
                    pdf_x + label_half_width + 2,
                    label_top_y + label_half_height + 2,
                )
            )

            c.setStrokeColor(HexColor("#dc2626"))
            c.setFillColor(HexColor("#dc2626"))
            c.setFont("Helvetica-Bold", font_size)
            c.drawString(pdf_x - text_width / 2, pdf_y - 2.2, label)

            c.setStrokeColor(HexColor("#dc2626"))
            source_box = expanded_horizontal_source_box(candidate, occupied_by_page.get(page_index, []))
            line_start_x, line_start_y = line_start_outside_source_box(source_box, (pdf_x, pdf_y), height)
            vec_x = line_start_x - pdf_x
            vec_y = line_start_y - pdf_y
            vec_len = (vec_x * vec_x + vec_y * vec_y) ** 0.5 or 1.0
            if vec_len > 1.5:
                label_edge_x = pdf_x + (vec_x / vec_len) * label_half_width
                label_edge_y = pdf_y + (vec_y / vec_len) * label_half_height
                c.line(line_start_x, line_start_y, label_edge_x, label_edge_y)

        c.save()

        overlay_reader = PdfReader(str(overlay_path))
        page.merge_page(overlay_reader.pages[0])
        writer.add_page(page)

    with output_pdf.open("wb") as handle:
        writer.write(handle)

    shutil.rmtree(tmp_dir, ignore_errors=True)


def load_candidates(path: Path) -> list[DimensionCandidate]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    candidates: list[DimensionCandidate] = []
    for item in payload:
        candidates.append(
            DimensionCandidate(
                number=int(item["number"]),
                page=int(item["page"]),
                text=str(item.get("text", "")),
                x=float(item["x"]),
                y=float(item["y"]),
                width=float(item.get("width", 0)),
                height=float(item.get("height", 0)),
                confidence=float(item.get("confidence", 0)),
                reason=str(item.get("reason", "manual")),
            )
        )
    candidates.sort(key=lambda item: (item.page, item.number))
    return candidates


def generate_tolerance_workbook(
    candidates: list[DimensionCandidate],
    output_xlsx: Path,
    job: dict[str, Any] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspeccion Final"

    metadata = job or {}
    dark_fill = PatternFill("solid", fgColor="111827")
    thin = Side(style="thin", color="000000")
    orange = Side(style="medium", color="F97316")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    check_border = Border(left=orange, right=orange, top=orange, bottom=orange)

    ws.merge_cells("A1:B5")
    logo_path = Path(__file__).resolve().parents[1] / "assets" / "smart-tool-logo.png"
    if logo_path.exists():
        logo = ExcelImage(str(logo_path))
        logo.width = 150
        logo.height = 72
        ws.add_image(logo, "A1")
    else:
        ws["A1"] = "SMART\nTOOL"
        ws["A1"].font = Font(bold=True, size=24, color="6B7280")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("C1:H1")
    ws["C1"] = "INSPECCION FINAL DE PRODUCTO"
    ws["C1"].font = Font(bold=True, size=16, name="Times New Roman")
    ws["C1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("I1:M1")
    ws["I1"] = "SR-P-19-02 Rev 04 Emision: 30/07/25"
    ws["I1"].alignment = Alignment(horizontal="center")

    form_cells = [
        ("C3", "Cliente:"), ("D3", metadata.get("client", "")),
        ("I3", "#F:"), ("J3", ""),
        ("L3", "PO#"), ("M3", ""),
        ("C5", "No. Parte:"), ("D5", metadata.get("part_number") or metadata.get("drawing_number", "")),
        ("I5", "Rev:"), ("J5", metadata.get("revision", "")),
    ]
    for cell_ref, value in form_cells:
        ws[cell_ref] = value
        ws[cell_ref].font = Font(bold=cell_ref[0] in {"C", "I", "L"})
    for merged in ["D3:H3", "J3:K3", "M3:M3", "D5:H5", "J5:K5"]:
        ws.merge_cells(merged)
    for row in [3, 5]:
        for col in range(4, 14):
            ws.cell(row=row, column=col).border = Border(bottom=thin)

    headers = [
        "ITEM",
        "NOMINAL",
        "INSTRUMENT",
        "Tol +",
        "Tol -",
        "PIEZA 1",
        "PIEZA 2",
        "PIEZA 3",
        "PIEZA 4",
        "PIEZA 5",
        "PIEZA 6",
        "PIEZA 7",
    ]
    header_row = 7
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = dark_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    tolerance_candidates = [
        candidate
        for candidate in sorted(candidates, key=lambda item: (item.page, item.number))
        if not is_date_like_text(candidate.text)
    ]
    for row, candidate in enumerate(tolerance_candidates, start=header_row + 1):
        info = tolerance_info(candidate.text)
        is_thread = is_thread_callout(candidate.text)
        values = [
            candidate.number,
            inspection_nominal_text(candidate.text),
            "",
            "" if is_thread else info["tol_plus"],
            "" if is_thread else info["tol_minus"],
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in {4, 5} and value not in {None, ""}:
                cell.number_format = "0.0000"
            if col in {1, 4, 5} or col >= 6:
                cell.alignment = Alignment(horizontal="center")
            if col == 2:
                cell.alignment = Alignment(horizontal="left")

    widths = [8, 18, 16, 12, 12, 14, 14, 14, 14, 14, 14, 14]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in range(header_row + 1, header_row + len(tolerance_candidates) + 1):
        ws.row_dimensions[row].height = 20
    ws.freeze_panes = "A8"
    last_data_row = max(header_row + 1, header_row + len(tolerance_candidates))
    ws.auto_filter.ref = f"A{header_row}:L{last_data_row}"

    footer_row = last_data_row + 1
    for col in range(1, 13):
        ws.cell(row=footer_row, column=col).border = border
        ws.cell(row=footer_row + 1, column=col).border = border
        ws.cell(row=footer_row + 2, column=col).border = border

    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4)
    ws.cell(row=footer_row, column=1, value="Realizado por:").font = Font(bold=True)
    ws.merge_cells(start_row=footer_row, start_column=5, end_row=footer_row, end_column=6)
    ws.cell(row=footer_row, column=5, value="Fecha:").font = Font(bold=True)
    ws.merge_cells(start_row=footer_row, start_column=7, end_row=footer_row, end_column=8)
    ws.cell(row=footer_row, column=7, value="Piezas Marcadas:")
    ws.cell(row=footer_row, column=9).border = check_border
    ws.merge_cells(start_row=footer_row, start_column=10, end_row=footer_row, end_column=11)
    ws.cell(row=footer_row, column=10, value="Flash Chroming:")
    ws.cell(row=footer_row, column=12).border = check_border

    ws.merge_cells(start_row=footer_row + 1, start_column=7, end_row=footer_row + 1, end_column=8)
    ws.cell(row=footer_row + 1, column=7, value="Material OK:")
    ws.cell(row=footer_row + 1, column=9).border = check_border
    ws.merge_cells(start_row=footer_row + 1, start_column=10, end_row=footer_row + 1, end_column=11)
    ws.cell(row=footer_row + 1, column=10, value="Heat Treatment:")
    ws.cell(row=footer_row + 1, column=12).border = check_border

    ws.merge_cells(start_row=footer_row + 2, start_column=1, end_row=footer_row + 2, end_column=6)
    ws.cell(row=footer_row + 2, column=1, value="Detalles de Inspeccion:").font = Font(bold=True)
    ws.merge_cells(start_row=footer_row + 2, start_column=7, end_row=footer_row + 2, end_column=8)
    ws.cell(row=footer_row + 2, column=7, value="Anodized:")
    ws.cell(row=footer_row + 2, column=9).border = check_border

    for row in range(footer_row, footer_row + 3):
        for col in range(1, 13):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center")

    rules = wb.create_sheet("Reglas")
    rules["A1"] = "Reglas de tolerancia general"
    rules["A1"].font = Font(bold=True, size=13)
    rules.append(["Decimales", "Tolerancia"])
    rules.append([2, 0.01])
    rules.append([3, 0.005])
    rules.append([4, 0.001])
    rules.append(["Prioridad", "Si la cota trae tolerancia explicita, se usa esa tolerancia antes de la regla general."])
    rules.append(["Roscas", "Se incluyen llamadas tipo 4-40 UNC, 6-32 UNC, etc.; no se les asigna tolerancia."])
    rules.append(["Radios", "Las cotas tipo R.063 se calculan usando el valor despues de R."])
    rules.append(["Fechas", "No se incluyen textos con formato de fecha como 10/22/2025."])
    for col in range(1, 3):
        rules.column_dimensions[get_column_letter(col)].width = 28

    wb.save(output_xlsx)


def init_history(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            create table if not exists jobs (
                id text primary key,
                client text,
                part_number text,
                drawing_number text,
                revision text,
                analysis_strategy text,
                source_hash text not null,
                original_pdf text not null,
                numbered_pdf text not null,
                candidates_json text not null,
                created_at text not null
            )
            """
        )
        conn.execute(
            "create index if not exists idx_jobs_lookup on jobs(client, part_number, drawing_number, revision)"
        )
        columns = {row[1] for row in conn.execute("pragma table_info(jobs)")}
        if "analysis_strategy" not in columns:
            try:
                conn.execute("alter table jobs add column analysis_strategy text")
            except sqlite3.OperationalError as exc:
                if "readonly" not in str(exc).lower():
                    raise


def save_history(db_path: Path, job: dict[str, Any]) -> None:
    with sqlite3.connect(db_path) as conn:
        columns = {row[1] for row in conn.execute("pragma table_info(jobs)")}
        if "analysis_strategy" not in columns:
            conn.execute(
                """
                insert or replace into jobs (
                    id, client, part_number, drawing_number, revision, source_hash,
                    original_pdf, numbered_pdf, candidates_json, created_at
                ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    job["id"],
                    job.get("client"),
                    job.get("part_number"),
                    job.get("drawing_number"),
                    job.get("revision"),
                    job["source_hash"],
                    job["original_pdf"],
                    job["numbered_pdf"],
                    job["candidates_json"],
                    job["created_at"],
                ),
            )
            return

        conn.execute(
            """
            insert or replace into jobs (
                id, client, part_number, drawing_number, revision, analysis_strategy, source_hash,
                original_pdf, numbered_pdf, candidates_json, created_at
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job["id"],
                job.get("client"),
                job.get("part_number"),
                job.get("drawing_number"),
                job.get("revision"),
                job.get("analysis_strategy"),
                job["source_hash"],
                job["original_pdf"],
                job["numbered_pdf"],
                job["candidates_json"],
                job["created_at"],
            ),
        )


def analyze_command(args: argparse.Namespace) -> None:
    input_pdf = Path(args.input_pdf).resolve()
    if not input_pdf.exists():
        raise SystemExit(f"PDF not found: {input_pdf}")

    storage = Path(args.storage).resolve()
    source_hash = sha256_file(input_pdf)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    identity = "-".join(
        part for part in [args.client, args.part_number, args.drawing_number, args.revision] if part
    )
    safe_identity = re.sub(r"[^A-Za-z0-9_.-]+", "-", identity).strip("-") or "plano"
    job_id = f"{safe_identity}-{stamp}-{source_hash[:8]}"
    job_dir = storage / "jobs" / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    original_pdf = job_dir / "original.pdf"
    candidates_json = job_dir / "candidates.json"
    numbered_pdf = job_dir / "numbered.pdf"
    tolerances_xlsx = job_dir / "tolerancias.xlsx"
    job_json = job_dir / "job.json"

    shutil.copy2(input_pdf, original_pdf)
    candidates, analysis_strategy, used_ocr = extract_candidates(
        original_pdf,
        include_tables=args.include_tables,
        strategy=args.strategy,
    )
    draw_numbered_overlay(original_pdf, candidates, numbered_pdf)

    candidates_payload = [asdict(candidate) for candidate in candidates]
    candidates_json.write_text(json.dumps(candidates_payload, indent=2), encoding="utf-8")

    job = {
        "id": job_id,
        "used_ocr": used_ocr,
        "analysis_strategy": analysis_strategy,
        "client": args.client,
        "part_number": args.part_number,
        "drawing_number": args.drawing_number,
        "revision": args.revision,
        "source_hash": source_hash,
        "original_pdf": str(original_pdf),
        "numbered_pdf": str(numbered_pdf),
        "candidates_json": str(candidates_json),
        "tolerances_xlsx": str(tolerances_xlsx),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "candidate_count": len(candidates),
    }
    generate_tolerance_workbook(candidates, tolerances_xlsx, job)
    job_json.write_text(json.dumps(job, indent=2), encoding="utf-8")

    db_path = storage / "history.sqlite"
    init_history(db_path)
    save_history(db_path, job)

    print(json.dumps(job, indent=2))


def search_command(args: argparse.Namespace) -> None:
    db_path = Path(args.storage).resolve() / "history.sqlite"
    init_history(db_path)
    params = {
        "client": args.client,
        "part_number": args.part_number,
        "drawing_number": args.drawing_number,
        "revision": args.revision,
    }
    clauses = []
    values = []
    for key, value in params.items():
        if value:
            clauses.append(f"{key} like ?")
            values.append(f"%{value}%")

    sql = "select * from jobs"
    if clauses:
        sql += " where " + " and ".join(clauses)
    sql += " order by created_at desc limit ?"
    values.append(args.limit)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = [dict(row) for row in conn.execute(sql, values)]

    print(json.dumps(rows, indent=2))


def render_command(args: argparse.Namespace) -> None:
    original_pdf = Path(args.original_pdf).resolve()
    candidates_json = Path(args.candidates_json).resolve()
    output_pdf = Path(args.output_pdf).resolve()
    if not original_pdf.exists():
        raise SystemExit(f"PDF not found: {original_pdf}")
    if not candidates_json.exists():
        raise SystemExit(f"Candidates JSON not found: {candidates_json}")

    output_pdf.parent.mkdir(parents=True, exist_ok=True)
    draw_numbered_overlay(original_pdf, load_candidates(candidates_json), output_pdf)
    print(json.dumps({"output_pdf": str(output_pdf)}, indent=2))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Detecta y numera cotas en planos PDF.")
    parser.add_argument("--storage", default="storage", help="Carpeta de historial.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    analyze = subparsers.add_parser("analyze", help="Analiza y numera un PDF.")
    analyze.add_argument("input_pdf")
    analyze.add_argument("--client", default="")
    analyze.add_argument("--part-number", default="")
    analyze.add_argument("--drawing-number", default="")
    analyze.add_argument("--revision", default="")
    analyze.add_argument(
        "--include-tables",
        action="store_true",
        help="Incluye numeros dentro de tablas. Por defecto se excluyen.",
    )
    analyze.add_argument(
        "--strategy",
        choices=[DEFAULT_ANALYSIS_STRATEGY, *ANALYSIS_PROFILES.keys()],
        default=DEFAULT_ANALYSIS_STRATEGY,
        help="Estrategia de deteccion: auto prueba varios perfiles; las otras fuerzan un perfil.",
    )
    analyze.set_defaults(func=analyze_command)

    search = subparsers.add_parser("search", help="Busca trabajos guardados.")
    search.add_argument("--client", default="")
    search.add_argument("--part-number", default="")
    search.add_argument("--drawing-number", default="")
    search.add_argument("--revision", default="")
    search.add_argument("--limit", type=int, default=20)
    search.set_defaults(func=search_command)

    render = subparsers.add_parser("render", help="Regenera un PDF numerado desde candidatos editados.")
    render.add_argument("original_pdf")
    render.add_argument("candidates_json")
    render.add_argument("output_pdf")
    render.set_defaults(func=render_command)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
